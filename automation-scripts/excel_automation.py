"""
Excel/Google Sheets automation template.
Common freelance task — sell for $30-100 on Fiverr.
Customize per client task.
"""

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference


def create_sales_report(data: list[dict], output_path: str = "report.xlsx"):
    """
    Creates a formatted sales report from raw data.
    data example: [{"product": "Item A", "sales": 150, "revenue": 4500.0}, ...]
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales Report"

    # Styles
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    headers = ["#", "Product", "Units Sold", "Revenue ($)", "Avg Price ($)"]
    col_widths = [5, 30, 15, 15, 15]

    # Write headers
    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 25

    # Write data rows
    for row_idx, item in enumerate(data, 2):
        avg_price = item["revenue"] / item["sales"] if item["sales"] > 0 else 0
        row_data = [row_idx - 1, item["product"], item["sales"], item["revenue"], round(avg_price, 2)]

        fill_color = "F8FAFC" if row_idx % 2 == 0 else "FFFFFF"
        row_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = row_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center")

    # Totals row
    total_row = len(data) + 2
    ws.cell(row=total_row, column=2, value="TOTAL").font = Font(bold=True)
    ws.cell(row=total_row, column=3, value=sum(d["sales"] for d in data)).font = Font(bold=True)
    ws.cell(row=total_row, column=4, value=sum(d["revenue"] for d in data)).font = Font(bold=True)

    # Add bar chart
    chart = BarChart()
    chart.title = "Revenue by Product"
    chart.y_axis.title = "Revenue ($)"
    chart.x_axis.title = "Product"

    data_ref = Reference(ws, min_col=4, min_row=1, max_row=len(data) + 1)
    cats = Reference(ws, min_col=2, min_row=2, max_row=len(data) + 1)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats)
    chart.shape = 4
    ws.add_chart(chart, f"A{total_row + 2}")

    wb.save(output_path)
    print(f"Report saved: {output_path}")


if __name__ == "__main__":
    sample_data = [
        {"product": "Product A", "sales": 245, "revenue": 12250.0},
        {"product": "Product B", "sales": 189, "revenue": 9450.0},
        {"product": "Product C", "sales": 312, "revenue": 15600.0},
        {"product": "Product D", "sales": 98, "revenue": 4900.0},
        {"product": "Product E", "sales": 421, "revenue": 21050.0},
    ]
    create_sales_report(sample_data, "sales_report.xlsx")
